# 导入所需模块
import openpyxl
import os

# 定义读取excel表格函数
def read_excel(file_path, sheet_name):
    # 打开excel文件
    wb = openpyxl.load_workbook(file_path)
    # 选择sheet
    sheet = wb[sheet_name]
    # 获取最大行数和列数
    max_row = sheet.max_row
    max_column = sheet.max_column
    # 定义一个空列表，用于存储每一行的数据
    data_list = []
    # 遍历每一行
    for row in range(1, max_row+1):
        # 定义一个空列表，用于存储每一列的数据
        row_list = []
        # 遍历每一列
        for column in range(1, max_column+1):
            # 获取单元格的值
            cell_value = sheet.cell(row=row, column=column).value
            # 将单元格的值添加到行列表中
            row_list.append(cell_value)
        # 将行列表添加到数据列表中
        data_list.append(row_list)
    # 返回数据列表
    return data_list

# 定义替换函数
def replace_txt(file_path, data_list):
    # 打开txt文件
    with open(file_path, 'r', encoding='utf-8') as f:
        # 读取txt文件中的内容
        content = f.read()
    # 遍历数据列表
    for data in data_list:
        # 获取需要替换的值和替换后的值
        old_value = data[0]
        new_value = data[1]
        # 判断需要替换的值是否存在于txt文件中
        if old_value in content:
            # 替换txt文件中的内容
            content = content.replace(old_value, new_value)
    # 将替换后的内容写入新的txt文件中
    with open('d:/test/x.txt', 'w', encoding='utf-8') as f:
        f.write(content)

# 读取sheet1
data_list = read_excel('excel文件路径', 'sheet1')
# 替换x.txt中的值
replace_txt('d:/test/x.txt', data_list)

# 读取sheet2
data_list = read_excel('excel文件路径', 'sheet2')
# 替换y.txt中的值
replace_txt('d:/test/y.txt', data_list)

# 读取sheet3
data_list = read_excel('excel文件路径', 'sheet3')
# 将数据写入z.txt中
with open('d:/test/z.txt', 'w', encoding='utf-8') as f:
    for data in data_list:
        f.write('\t'.join(data) + '\n')